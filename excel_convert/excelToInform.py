from openpyxl import load_workbook,Workbook


def removeSlash(str):
    ans = str.split('/')
    return ans
    

load_wb = load_workbook('excelConvert/testInput.xlsx', data_only=True)
load_ws = load_wb['시트1']
write_wb = Workbook()
write_ws = write_wb.create_sheet('참석자 시트')
write_ws = write_wb.active
i_list = [1,3]

for i in range(2,15):
    for j in i_list:
        people = load_ws.cell(i,j).value
        if people:
            people = removeSlash(people)
            write_ws.append(people)
write_wb.save("excelConvert/testOutput.xlsx")
