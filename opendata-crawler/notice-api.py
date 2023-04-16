from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from enum import Enum
#TODO:excel에 데이터 저장
# 가져온 데이터를 기반으로 엑셀 파일 생성

class ColNameKey(Enum):
   
    TITLE = '데이터 제목'
    DETAIL = '상세설명'
    LINK = '바로가기 링크'
    FACILITY ='제공기관'
    MODIFYAT = '수정일'
    VIEW = '조회수'
    DOWNLOAD = '다운로드'
    FREQUENTDATA = '주기성 데이터'
    KEYWORD = '키워드'
    
   
write_wb = Workbook()
sheet = write_wb.active
sheet.title = "고용노동 데이터"
col_name = [e.value for e in ColNameKey]
# print(col_name)
for idx, name in enumerate(col_name):
    sheet.cell(row=1,column=idx+1,value=name)
def writeWorkbook(idx,key,value):
    try:
        sheet.cell(row=idx,column=col_name.index(ColNameKey(key).value)+1,value=value)
    except:
        sheet.cell(row=idx,column=col_name.index(key)+1,value=value)
        
# ChromeDriver 경로 설정
DRIVER_PATH = 'C:/Users/MinSeo/Downloads/chromedriver_win32/chromdriver.exe'
URL = 'https://www.moel.go.kr/info/publicdata/publicopen/list.do'
# ChromeDriver 실행 옵션 설정
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('disable-gpu')
# ServiceProcess 생성
service = Service(DRIVER_PATH)
# WebDriver 생성
driver = webdriver.Chrome(service=service, options=options)

# 웹 페이지 접속
driver.get(URL)

# iframe 요소 선택
iframe = driver.find_element(By.NAME,'frm')
# iframe 내부로 이동
driver.switch_to.frame(iframe)

# pagination 적용

index = 2
# 정보 요소 가져오기
for page in range(1,19):
    #TODO: pagination 미적용 오류 개선
    driver.execute_script(
        #    "$('#search-form-current-page').val(arguments[0]);",page
        "eventFnObj.fn_dTypeClick('FILE');updatePage(arguments[0])",page
    )
    element = driver.find_element(By.CSS_SELECTOR ,'#fileDataList div.result-list ul')
    #list로 가져오기
    data_list = element.find_elements(By.TAG_NAME , 'li')
    for data in data_list:
        detail_title_area = data.find_element(By.CSS_SELECTOR ,'dl')
        title = detail_title_area.find_element(By.CSS_SELECTOR ,'dt a span.title').text
        writeWorkbook(index,ColNameKey.TITLE,title)
        # display 속성을 변경하여 화면에 표시되도록 만듦
        # @media 태그로 인해서 바로 수정되지않아, viewport 크기를 변경함으로써
        # 수정될수있도록 함
        driver.execute_script(
            "Object.defineProperty(window, 'innerWidth', {value: 1025, configurable: true});var dd_list = document.querySelectorAll('.data-set-list .result-list ul li dl dd');for (var i = 0; i < dd_list.length; i++) {dd_list[i].style.display = 'block';}")
        details = detail_title_area.find_elements(By.TAG_NAME ,"dd")
        link = detail_title_area.find_element(By.CSS_SELECTOR ,'dt a').get_attribute('href')
        writeWorkbook(index,ColNameKey.LINK,link)
        print('링크 테스트'+link)
        for detail in details:
            detail_list = ''
            detail_list += detail.text
            writeWorkbook(index,ColNameKey.DETAIL,detail_list)
        info_data = data.find_elements(By.CSS_SELECTOR ,'div.info-data p')
        for info in info_data:
            span_title = info.find_element(By.CSS_SELECTOR ,'span.tit').text
            try:
                span_data = info.find_element(By.CSS_SELECTOR ,'span.data').text
            except:
                tmp_span_data = info.text.split(' ')[1].split(',')
                span_data = ','.join(map(str, tmp_span_data))
            writeWorkbook(index,ColNameKey(span_title).value,span_data)
        index += 1
    
        

# iframe 바깥으로 이동
driver.switch_to.default_content()
write_wb.save("./testOutput.xlsx")   

# 웹 브라우저 종료
driver.quit()

